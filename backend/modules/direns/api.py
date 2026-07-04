"""Module DirENS — génération de l'Excel financier officiel pré-rempli.

Remplit de façon 100 % automatique et déterministe les DEUX premiers onglets du
template officiel DirENS (bundlé dans assets/template_direns.xlsx) à partir des
données de l'app, au plus proche d'un compte de résultat associatif :

  - COLONNES = les clubs ACTIFS (entités internes ayant des dépenses sur la
    période), ordonnés comme dans l'app. Un club sans activité n'apparaît pas.
  - LIGNES = les CATÉGORIES OpenFlow, en respectant la HIÉRARCHIE : une catégorie
    parente devient un en-tête (gras), ses sous-catégories sont indentées en
    dessous. Les catégories feuilles (sans enfant) sont des lignes simples.
  - Les cellules sans information restent VIDES (jamais de 0).
  - Section financement : les catégories au solde POSITIF (recettes > dépenses).
    Soldes de trésorerie estimés.
  - Onglet 3 « Demande subventions » : laissé vierge.

Aucune configuration : tout est déduit des catégories et entités de chaque
transaction. Le bloc des dépenses (et celui des financements) est redimensionné
dynamiquement, les formules de totaux sont recalculées, le nom et l'année du
titre sont dérivés du mandat (exercice) choisi.

Règle de répartition UNIQUE, sans cas particulier : pour chaque (club, catégorie),
solde net = recettes − dépenses. Net positif → financement (sous le nom de la
catégorie) ; net négatif → dépense réelle. Aucune distinction subvention/activité,
aucun plan comptable, aucun mot-clé : une catégorie qui rapporte va en financement,
une catégorie qui coûte va en dépense.

Convention monétaire : amount = entier de centimes, positif. Sens porté par
from_entity_id -> to_entity_id (dépense = from interne ; recette = to interne).
"""
import io
import re
import sqlite3
import zipfile
from copy import copy
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response

from backend.core.auth import require_admin
from backend.core.balance import compute_entity_balance
from backend.core.config import load_config
from backend.core.database import get_conn

router = APIRouter(dependencies=[Depends(require_admin)])

PROJECT_ROOT = Path(__file__).parent.parent.parent.parent
CONFIG_PATH = PROJECT_ROOT / "config.yaml"
ASSETS_DIR = Path(__file__).parent / "assets"
TEMPLATE_PATH = ASSETS_DIR / "template_direns.xlsx"

# Disposition du template (lignes Excel absolues du modèle d'origine).
EXP_FIRST = 7          # première ligne du bloc dépenses
EXP_SLOTS = 26         # lignes de saisie dépenses dans le modèle (7..32)
TOTAL_ROW = 33         # TOTAL DEPENSES (modèle)
FIN_FIRST = 35         # première ligne de financement (modèle)
FIN_SLOTS = 3          # lignes de saisie financement (35..37)
FIN_TOTAL = 38         # TOTAL FINANCEMENT RECU (modèle)
DEP_TOTAL = 39         # TOTAL DEPENSES REELLES (rappel) (modèle)
SOLDE_PASS = 40        # SOLDE COMPTE BANCAIRE (Date passation)
SOLDE_TRES = 41        # SOLDE TRESORERIE (A date)
SOLDE_DATE = 42        # SOLDE COMPTE BANCAIRE (A date)
LABEL_NONE = "Non catégorisé"
PLACEHOLDER = "à compléter"   # cellules non déductibles des données (à saisir à la main)
HEADER_TPL_ROW = 7     # ligne modèle « ACHAT » (en-tête gras)
DATA_TPL_ROW = 8       # ligne modèle de saisie (normale)


# ───────────────────────── Accès aux données ──────────────────────────────

def _get_clubs(conn) -> list:
    """Toutes les entités internes (clubs), ordonnées comme dans l'app."""
    rows = conn.execute(
        "SELECT id, name FROM entities WHERE type = 'internal' ORDER BY position ASC, id ASC"
    ).fetchall()
    return [dict(r) for r in rows]


def _active_clubs(all_clubs: list, amounts: dict) -> list:
    """Clubs ayant au moins un montant non nul (sinon tous, pour ne pas vider l'onglet)."""
    active = {club_id for (club_id, _), cents in amounts.items() if cents}
    clubs = [c for c in all_clubs if c["id"] in active]
    return clubs or all_clubs


def _resolve_fy(conn, fy_id: int) -> dict:
    row = conn.execute(
        "SELECT id, name, start_date, end_date FROM fiscal_years WHERE id = ?", (fy_id,)
    ).fetchone()
    if row is None:
        raise HTTPException(404, f"Exercice {fy_id} introuvable")
    end = row["end_date"] or date.today().isoformat()
    return {"id": row["id"], "name": row["name"], "start": row["start_date"], "end": end}


def _year_label(start: str, end: str) -> str:
    """Année universitaire « AAAA-AAAA » dérivée des dates de l'exercice/mandat."""
    try:
        sy = int(start[:4])
        ey = int(end[:4]) if end else sy
    except Exception:
        return ""
    return f"{sy}-{ey}" if ey > sy else f"{sy}-{sy + 1}"


def _categories(conn) -> dict:
    return {r["id"]: dict(r) for r in conn.execute(
        "SELECT id, name, parent_id FROM categories"
    ).fetchall()}


def _split_net_and_financement(exp_amounts: dict, income_by_cell: dict) -> tuple:
    """Répartit dépenses et recettes par (club, catégorie) selon UNE seule règle, sans aucun
    cas particulier : solde net = recettes − dépenses.
          · net > 0 → financement, pour +net, sous le nom de la catégorie.
          · net < 0 → dépense réelle, pour −net (= dépenses − recettes).
          · net = 0 → rien à afficher.

    Retourne (net, financement) :
      - net {(club, cat): cents}  : dépense réelle nette, par club et catégorie.
      - financement {cat: cents}  : soldes positifs, agrégés par catégorie.

    Recettes et dépenses d'une même catégorie se compensent ; le bilan reste équilibré
    (financement − dépenses réelles = recettes − dépenses).
    """
    net: dict = {}
    financement: dict = {}
    for key in set(exp_amounts) | set(income_by_cell):
        _club, cat = key
        bal = income_by_cell.get(key, 0) - exp_amounts.get(key, 0)
        if bal > 0:
            financement[cat] = financement.get(cat, 0) + bal
        elif bal < 0:
            net[key] = -bal
    return net, financement


def _get_realized_amounts(conn, start: str, end: str, club_ids: list, direction: str) -> dict:
    """{(club_id, category_id): cents} des dépenses ('expense') ou recettes ('income') réelles
    d'un club : flux avec l'externe uniquement (virements internes exclus des deux côtés)."""
    if not club_ids:
        return {}
    ph = ",".join("?" * len(club_ids))
    side, other = ("from_entity_id", "to_entity_id") if direction == "expense" else ("to_entity_id", "from_entity_id")
    rows = conn.execute(
        f"""SELECT {side} AS eid, category_id AS cid, SUM(amount) AS total
            FROM transactions
            WHERE date BETWEEN ? AND ?
              AND {side} IN ({ph})
              AND ({other} IS NULL OR {other} NOT IN ({ph}))
            GROUP BY {side}, category_id""",
        [start, end] + club_ids + club_ids,
    ).fetchall()
    return {(r["eid"], r["cid"]): r["total"] for r in rows if r["total"]}


def _get_budget_amounts(conn, fy_id: int, club_ids: list, direction: str) -> dict:
    """{(club_id, category_id): cents} des allocations budgétaires d'un sens ('expense'/'income')."""
    if not club_ids:
        return {}
    ph = ",".join("?" * len(club_ids))
    try:
        rows = conn.execute(
            f"""SELECT entity_id AS eid, category_id AS cid, SUM(amount) AS total
                FROM budget_allocations
                WHERE fiscal_year_id = ? AND direction = ? AND entity_id IN ({ph})
                GROUP BY entity_id, category_id""",
            [fy_id, direction] + club_ids,
        ).fetchall()
    except sqlite3.OperationalError:
        return {}
    return {(r["eid"], r["cid"]): r["total"] for r in rows if r["total"]}


def _category_rows(conn, amounts: dict, clubs: list) -> list:
    """Construit les lignes hiérarchiques du bloc dépenses.

    Chaque ligne est un dict : {label, level, header, euros}
      - parent ayant des enfants -> ligne en-tête (header=True, level=0, pas de valeurs)
      - sous-catégorie            -> ligne indentée (level=1) avec ses montants
      - catégorie feuille racine  -> ligne simple (level=0) avec ses montants
    Triées par nom (en-têtes/feuilles mélangés alphabétiquement, enfants sous leur parent).
    """
    cats = _categories(conn)
    id_to_idx = {c["id"]: i for i, c in enumerate(clubs)}
    per_cat: dict = {}
    for (club_id, cat_id), cents in amounts.items():
        idx = id_to_idx.get(club_id)
        if idx is None:
            continue
        per_cat.setdefault(cat_id, {})
        per_cat[cat_id][idx] = per_cat[cat_id].get(idx, 0) + cents

    children: dict = {}
    for cid, c in cats.items():
        if c["parent_id"] is not None:
            children.setdefault(c["parent_id"], []).append(cid)

    def total(cid):
        return sum(per_cat.get(cid, {}).values())

    def euros(cid):
        return {idx: round(c / 100, 2) for idx, c in per_cat.get(cid, {}).items() if c}

    def name(cid):
        return cats.get(cid, {}).get("name") or LABEL_NONE

    rows = []
    top_ids = sorted(
        [cid for cid, c in cats.items() if c["parent_id"] is None],
        key=lambda cid: name(cid).lower(),
    )
    for cid in top_ids:
        kids = children.get(cid, [])
        if kids:
            active_kids = sorted([k for k in kids if total(k) != 0], key=lambda k: name(k).lower())
            own = total(cid)
            if not active_kids and own == 0:
                continue
            rows.append({"label": name(cid), "level": 0, "header": True, "euros": {}})
            if own != 0:  # dépense directe sur le parent (rare) : ligne indentée dédiée
                rows.append({"label": f"{name(cid)} (divers)", "level": 1, "header": False, "euros": euros(cid)})
            for k in active_kids:
                rows.append({"label": name(k), "level": 1, "header": False, "euros": euros(k)})
        else:
            if total(cid) == 0:
                continue
            rows.append({"label": name(cid), "level": 0, "header": False, "euros": euros(cid)})

    if None in per_cat and total(None) != 0:
        rows.append({"label": LABEL_NONE, "level": 0, "header": False, "euros": euros(None)})
    return rows


def _financement_rows(conn, financement: dict) -> list:
    """Lignes de la section financement [(label, euros)] : toutes les catégories au solde net
    positif (recettes > dépenses), sous leur propre nom, triées par ordre alphabétique
    (cf. _split_net_and_financement). Aucune distinction catégorie-ressource / activité."""
    names = _categories(conn)
    out = []
    for cid, cents in financement.items():
        if not cents:
            continue
        label = (names.get(cid, {}).get("name") if cid is not None else None) or LABEL_NONE
        out.append((label, round(cents / 100, 2)))
    out.sort(key=lambda x: (x[0] == LABEL_NONE, x[0].lower()))
    return out


def _opening_explicit(conn, fy_id: int, club_ids: list) -> Optional[float]:
    """Solde d'ouverture SAISI (fiscal_year_opening_balances) en euros, ou None."""
    if not club_ids:
        return None
    ph = ",".join("?" * len(club_ids))
    try:
        rows = conn.execute(
            f"SELECT amount FROM fiscal_year_opening_balances "
            f"WHERE fiscal_year_id = ? AND entity_id IN ({ph})",
            [fy_id] + club_ids,
        ).fetchall()
    except sqlite3.OperationalError:
        return None
    if not rows:
        return None
    return round(sum(r["amount"] for r in rows) / 100, 2)


def _get_current_total(conn, club_ids: list) -> float:
    """Trésorerie courante estimée = somme des soldes consolidés des clubs (euros)."""
    total = sum(compute_entity_balance(conn, eid)["balance"] for eid in club_ids)
    return round(total / 100, 2)


def _resolve_assoc_name(provided: str) -> str:
    if provided and provided.strip():
        return provided.strip()
    try:
        cfg = load_config(str(CONFIG_PATH))
        return cfg.entity.name or "Association"
    except Exception:
        return "Association"


# ───────────────────────── Génération du fichier ──────────────────────────

def _safe_title(title: str) -> str:
    for ch in "[]:*?/\\":
        title = title.replace(ch, "-")
    return title[:31]


def _capture_row_style(ws, row: int, total_col: int) -> dict:
    out = {}
    for c in range(1, total_col + 1):
        cell = ws.cell(row=row, column=c)
        if cell.has_style:
            out[c] = copy(cell._style)
    return out


def _apply_row_style(ws, row: int, styles: dict):
    for c, style in styles.items():
        ws.cell(row=row, column=c)._style = copy(style)


def _widen_sheet(ws, extra: int, total_col: int):
    """Insère `extra` colonnes (clubs > 4) en recopiant le style de la colonne E."""
    from openpyxl.utils import get_column_letter

    ws.insert_cols(6, extra)
    for r in range(1, 43):
        src = ws.cell(row=r, column=5)
        for c in range(6, 6 + extra):
            dst = ws.cell(row=r, column=c)
            if src.has_style:
                dst._style = copy(src._style)
    e_width = ws.column_dimensions["E"].width
    if e_width:
        for c in range(6, 6 + extra):
            ws.column_dimensions[get_column_letter(c)].width = e_width
    tcl = get_column_letter(total_col)
    for row in (1, 3):
        existing = [str(m) for m in ws.merged_cells.ranges if m.min_row == row and m.max_row == row]
        for m in existing:
            ws.unmerge_cells(m)
        ws.merge_cells(f"A{row}:{tcl}{row}")


def _layout(ws, n_clubs: int):
    """Prépare les colonnes (insertion si >4 clubs) et renvoie (last_club_col, total_col)."""
    if n_clubs <= 4:
        return 5, 6
    extra = n_clubs - 4
    last_club_col, total_col = n_clubs + 1, n_clubs + 2
    _widen_sheet(ws, extra, total_col)
    return last_club_col, total_col


def _fill_sheet(ws, clubs, last_club_col, total_col, expense_rows, income_rows,
                with_financing, total_label, year_label="", opening=None, current=None):
    """Écrit en-têtes clubs, lignes catégories hiérarchiques, totaux et (si bilan)
    la section financement + soldes. Redimensionne les blocs au nombre de lignes."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    # Styles de référence capturés AVANT toute insertion/suppression de lignes.
    data_style = _capture_row_style(ws, DATA_TPL_ROW, total_col)
    header_style = _capture_row_style(ws, HEADER_TPL_ROW, total_col)
    fin_style = _capture_row_style(ws, FIN_FIRST, total_col) if with_financing else {}

    n = len(clubs)
    for i, club in enumerate(clubs):
        ws.cell(row=5, column=2 + i).value = club["name"]

    # ── Redimensionnement du bloc dépenses ──
    e = len(expense_rows)
    if e < EXP_SLOTS:
        ws.delete_rows(EXP_FIRST + e, EXP_SLOTS - e)
    elif e > EXP_SLOTS:
        ws.insert_rows(TOTAL_ROW, e - EXP_SLOTS)
    delta_exp = e - EXP_SLOTS
    total_row = TOTAL_ROW + delta_exp

    fin_first = FIN_FIRST + delta_exp
    fin_total = FIN_TOTAL + delta_exp
    dep_total = DEP_TOTAL + delta_exp
    solde_pass = SOLDE_PASS + delta_exp
    solde_tres = SOLDE_TRES + delta_exp
    solde_date = SOLDE_DATE + delta_exp

    if with_financing:
        i_count = len(income_rows)
        if i_count < FIN_SLOTS:
            ws.delete_rows(fin_first + i_count, FIN_SLOTS - i_count)
        elif i_count > FIN_SLOTS:
            ws.insert_rows(fin_total, i_count - FIN_SLOTS)
        delta_fin = i_count - FIN_SLOTS
        fin_total += delta_fin
        dep_total += delta_fin
        solde_pass += delta_fin
        solde_tres += delta_fin
        solde_date += delta_fin

    first_cl = get_column_letter(2)
    last_cl = get_column_letter(last_club_col)
    total_cl = get_column_letter(total_col)

    # Valeurs calculées des cellules-formules, à injecter en cache dans le .xlsx final
    # (sinon les totaux s'affichent vides tant que le lecteur ne recalcule pas). {coord: valeur}.
    cache: dict = {}

    # ── Écriture des lignes catégories (hiérarchie) ──
    for k, row in enumerate(expense_rows):
        r = EXP_FIRST + k
        _apply_row_style(ws, r, header_style if row["header"] else data_style)
        a = ws.cell(row=r, column=1)
        a.value = row["label"]
        if row["level"] > 0:
            cur = a.alignment
            a.alignment = Alignment(horizontal="left", vertical=cur.vertical, indent=row["level"])
        if not row["header"]:
            for idx, val in row["euros"].items():
                ws.cell(row=r, column=2 + idx).value = val
            ws.cell(row=r, column=total_col).value = f"=SUM({first_cl}{r}:{last_cl}{r})"
            cache[f"{total_cl}{r}"] = round(sum(row["euros"].values()), 2)  # somme horizontale

    # ── Ligne TOTAL dépenses ── (sommes verticales par club + total général)
    ws.cell(row=total_row, column=1).value = total_label
    col_totals = {
        i: sum(row["euros"].get(i, 0) for row in expense_rows if not row["header"])
        for i in range(n)
    }
    grand_total = round(sum(col_totals.values()), 2)
    if e > 0:
        for i in range(n):
            cl = get_column_letter(2 + i)
            ws.cell(row=total_row, column=2 + i).value = f"=SUM({cl}{EXP_FIRST}:{cl}{total_row - 1})"
            cache[f"{cl}{total_row}"] = round(col_totals[i], 2)
        ws.cell(row=total_row, column=total_col).value = f"=SUM({first_cl}{total_row}:{last_cl}{total_row})"
        cache[f"{total_cl}{total_row}"] = grand_total
    else:
        ws.cell(row=total_row, column=total_col).value = 0

    # ── Section financement + soldes (bilan uniquement) ──
    if with_financing:
        i_count = len(income_rows)
        for k, (label, val) in enumerate(income_rows):
            r = fin_first + k
            _apply_row_style(ws, r, fin_style)
            ws.cell(row=r, column=1).value = label
            ws.cell(row=r, column=2).value = val
        ws.cell(row=fin_total, column=1).value = f"TOTAL FINANCEMENT RECU {year_label}".rstrip()
        ws.cell(row=fin_total, column=2).value = (
            f"=SUM(B{fin_first}:B{fin_first + i_count - 1})" if i_count > 0 else 0
        )
        if i_count > 0:
            cache[f"B{fin_total}"] = round(sum(val for _, val in income_rows), 2)
        ws.cell(row=dep_total, column=1).value = (
            f"TOTAL DEPENSES REELLES {year_label} (cf tableau ci-dessus)".replace("  ", " ")
        )
        ws.cell(row=dep_total, column=2).value = f"={total_cl}{total_row}"
        cache[f"B{dep_total}"] = grand_total
        # Solde bancaire à la passation : valeur saisie si disponible, sinon placeholder.
        ws.cell(row=solde_pass, column=2).value = opening if opening is not None else PLACEHOLDER
        # Solde trésorerie (à date) : estimation à partir des données de l'app.
        if current is not None:
            ws.cell(row=solde_tres, column=2).value = current
        # Solde compte bancaire (à date) : non déductible (l'app ne distingue pas encore
        # compte courant / Livret A / caisse physique) -> placeholder.
        ws.cell(row=solde_date, column=2).value = PLACEHOLDER

    return cache


def _patch_sheet_xml(xml_bytes: bytes, cache: dict) -> bytes:
    """Insère la valeur calculée `<v>` dans chaque cellule-formule listée dans `cache`
    ({coord: valeur}), en UNE seule passe regex, sans re-sérialiser le reste du XML (donc
    aucun risque de mélanger les préfixes de namespaces OOXML)."""
    if not cache:
        return xml_bytes
    num_by_coord = {coord: repr(round(val, 2)) for coord, val in cache.items()}
    # Un seul motif pour toutes les cellules ciblées (chaque coord est unique dans la feuille).
    cell_re = re.compile(
        r'(<c\b[^>]*\br="(' + "|".join(re.escape(c) for c in num_by_coord) + r')"[^>]*>)(.*?)(</c>)',
        re.DOTALL)

    def repl(m):
        head, coord, body, tail = m.group(1), m.group(2), m.group(3), m.group(4)
        if "<f" not in body:  # uniquement les cellules à formule
            return m.group(0)
        vtag = f"<v>{num_by_coord[coord]}</v>"
        # openpyxl écrit normalement un <v></v> VIDE après la formule : on le remplit ;
        # sinon (aucun <v>) on l'ajoute.
        body, replaced = re.subn(r"<v\s*/>|<v>.*?</v>", vtag, body, count=1, flags=re.DOTALL)
        if not replaced:
            body = f"{body}{vtag}"
        return f"{head}{body}{tail}"

    return cell_re.sub(repl, xml_bytes.decode("utf-8")).encode("utf-8")


def _inject_cached_values(xlsx_bytes: bytes, caches: dict) -> bytes:
    """Réécrit le paquet .xlsx en stockant, à côté de chaque formule de total, sa VALEUR
    calculée (cache `<v>`). Les totaux s'affichent alors dans tout lecteur sans recalcul, tout
    en restant des formules `=SUM(...)` éditables. `caches = {sheet_index: {coord: valeur}}` ;
    l'ordre des feuilles dans le zip (sheet1.xml, sheet2.xml…) suit `wb.worksheets`."""
    prefix, suffix = "xl/worksheets/sheet", ".xml"
    src = zipfile.ZipFile(io.BytesIO(xlsx_bytes))
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            name = item.filename
            if name.startswith(prefix) and name.endswith(suffix):
                try:
                    idx = int(name[len(prefix):-len(suffix)]) - 1
                except ValueError:
                    idx = None
                if cache := caches.get(idx):
                    data = _patch_sheet_xml(data, cache)
            dst.writestr(item, data)
    return out.getvalue()


def _build_excel(conn, bilan_fy: dict, budget_fy: Optional[dict], assoc_name: str) -> bytes:
    from openpyxl import load_workbook

    wb = load_workbook(str(TEMPLATE_PATH))
    all_clubs = _get_clubs(conn)
    all_ids = [c["id"] for c in all_clubs]

    # ── Onglet 1 : bilan financier (réalisé) ──
    # Règle unique par (club, catégorie) : net = recettes − dépenses ; net > 0 → financement,
    # net < 0 → dépense réelle. Aucune distinction subvention/activité (cf _split_net_and_financement).
    caches: dict = {}
    ws1 = wb.worksheets[0]
    exp_amounts = _get_realized_amounts(conn, bilan_fy["start"], bilan_fy["end"], all_ids, "expense")
    income_by_cell = _get_realized_amounts(conn, bilan_fy["start"], bilan_fy["end"], all_ids, "income")
    net_amounts, financement = _split_net_and_financement(exp_amounts, income_by_cell)
    clubs1 = _active_clubs(all_clubs, net_amounts)
    last1, total1 = _layout(ws1, len(clubs1))
    bilan_year = _year_label(bilan_fy["start"], bilan_fy["end"])
    caches[0] = _fill_sheet(
        ws1, clubs1, last1, total1,
        _category_rows(conn, net_amounts, clubs1),
        _financement_rows(conn, financement),
        with_financing=True, total_label="TOTAL DEPENSES REELLES", year_label=bilan_year,
        opening=_opening_explicit(conn, bilan_fy["id"], all_ids),
        current=_get_current_total(conn, all_ids),
    )
    ws1["A1"] = f"BILAN FINANCIER {bilan_year}"
    ws1["A3"] = assoc_name
    try:
        ws1.title = _safe_title(f"Bilan financier {bilan_year}")
    except Exception:
        pass

    # ── Onglet 2 : budget prévisionnel ──
    # Même règle nette que le bilan. Le template du prévisionnel n'a pas de section financement
    # (il s'arrête au TOTAL) : une catégorie budgétée bénéficiaire (recette > dépense, net positif
    # → `_bfin`) disparaît entièrement de l'onglet, y compris sa dépense prévue. Le TOTAL peut donc
    # être inférieur à la somme brute des dépenses prévues (limitation acceptée, faute de section).
    if budget_fy:
        ws2 = wb.worksheets[1]
        bexp_amounts = _get_budget_amounts(conn, budget_fy["id"], all_ids, "expense")
        bincome = _get_budget_amounts(conn, budget_fy["id"], all_ids, "income")
        bnet_amounts, _bfin = _split_net_and_financement(bexp_amounts, bincome)
        clubs2 = _active_clubs(all_clubs, bnet_amounts)
        last2, total2 = _layout(ws2, len(clubs2))
        budget_year = _year_label(budget_fy["start"], budget_fy["end"])
        caches[1] = _fill_sheet(
            ws2, clubs2, last2, total2,
            _category_rows(conn, bnet_amounts, clubs2), [],
            with_financing=False, total_label="TOTAL DEPENSES PREVISONNELLES (E)",
            year_label=budget_year,
        )
        ws2["A1"] = f"BUDGET PREVISIONNEL {budget_year}"
        ws2["A3"] = assoc_name
        try:
            ws2.title = _safe_title(f"Budget prévisionnel {budget_year}")
        except Exception:
            pass

    buf = io.BytesIO()
    wb.save(buf)
    return _inject_cached_values(buf.getvalue(), caches)


# ───────────────────────── Endpoint ────────────────────────────────────────

@router.get("/export")
def export_direns(
    bilan_fiscal_year_id: int,
    budget_fiscal_year_id: Optional[int] = None,
    assoc_name: str = "",
):
    """Génère le fichier Excel DirENS pré-rempli (onglets 1 et 2)."""
    if not TEMPLATE_PATH.exists():
        raise HTTPException(500, "Template DirENS introuvable dans le module (assets/template_direns.xlsx).")
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise HTTPException(500, "openpyxl non installé : impossible de générer le fichier Excel.")

    conn = get_conn()
    try:
        bilan_fy = _resolve_fy(conn, bilan_fiscal_year_id)
        budget_fy = _resolve_fy(conn, budget_fiscal_year_id) if budget_fiscal_year_id else None
        name = _resolve_assoc_name(assoc_name)
        try:
            data = _build_excel(conn, bilan_fy, budget_fy, name)
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(500, f"Erreur lors de la génération du fichier Excel : {e}")
    finally:
        conn.close()

    year = _year_label(bilan_fy["start"], bilan_fy["end"]) or bilan_fy["name"]
    safe = year.replace(" ", "_").replace("/", "-")
    filename = f"DirENS_{safe}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
