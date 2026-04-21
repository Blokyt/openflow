"""Parser for the BDA ENS Paris-Saclay compta Excel format.

Expected structure: sheet 'Suivi' with columns:
  Date | Motif | Catégorie | Dépense | Recette | Payeur | Remboursé | Facture
"""
from datetime import datetime

from .base import Parser, ParseResult, TransactionDraft


class BdaExcelParser(Parser):
    id = "bda_excel"
    name = "BDA (Excel)"
    description = "Format compta BDA avec onglet 'Suivi' (Date/Motif/Catégorie/Dépense/Recette/Payeur)"
    supported_extensions = [".xlsx", ".xls"]

    @staticmethod
    def detect(file_path: str, ext: str) -> float:
        if ext not in (".xlsx", ".xls"):
            return 0.0
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        except Exception:
            return 0.0

        score = 0.0
        # Must have a Suivi sheet
        if "Suivi" not in wb.sheetnames:
            wb.close()
            return 0.0
        score += 0.5

        # Check header row matches expected columns
        ws = wb["Suivi"]
        try:
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            wb.close()
            return score * 0.3
        wb.close()

        expected = {"date", "motif", "catégorie", "categorie", "dépense", "depense", "recette"}
        matched = sum(1 for h in header if h and str(h).strip().lower() in expected)
        if matched >= 4:
            score += 0.5
        elif matched >= 2:
            score += 0.25

        return min(score, 1.0)

    def parse(self, file_path: str) -> ParseResult:
        import openpyxl
        result = ParseResult(parser_id=self.id, parser_name=self.name, confidence=1.0)

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            result.errors.append(f"Impossible d'ouvrir le fichier: {e}")
            return result

        if "Suivi" not in wb.sheetnames:
            result.errors.append("Onglet 'Suivi' introuvable")
            return result

        ws = wb["Suivi"]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
            if len(row) < 5:
                continue
            date_val = row[0]
            label = row[1]
            cat = row[2]
            depense = row[3]
            recette = row[4]
            payeur = row[5] if len(row) > 5 else None
            rembourse = row[6] if len(row) > 6 else None
            facture = row[7] if len(row) > 7 else None

            if not date_val or (not depense and not recette):
                continue

            # Parse date
            if isinstance(date_val, datetime):
                date_str = date_val.strftime("%Y-%m-%d")
            elif isinstance(date_val, str):
                date_str = date_val[:10]
            else:
                result.warnings.append(f"Ligne {row_idx}: date invalide {date_val}")
                continue

            label_str = str(label or "").strip()
            cat_str = str(cat or "").strip()
            try:
                dep = float(depense) if depense else 0.0
                rec = float(recette) if recette else 0.0
            except (TypeError, ValueError):
                result.warnings.append(f"Ligne {row_idx}: montant invalide")
                continue

            amount = -dep if dep > 0 else rec
            if amount == 0:
                continue

            desc_parts = []
            reimbursement = None
            if payeur:
                p = str(payeur).strip()
                if p.lower() not in ("oui", "non", ""):
                    status = "pending"
                    if rembourse and str(rembourse).strip().lower() == "oui":
                        status = "reimbursed"
                    reimbursement = {"person_name": p, "status": status}
            if facture:
                f = str(facture).strip()
                if f:
                    desc_parts.append(f"Facture: {f}")

            result.transactions.append(TransactionDraft(
                date=date_str,
                label=label_str,
                amount=amount,
                description=" | ".join(desc_parts),
                category_hint=cat_str,
                raw={"row": row_idx, "payeur": str(payeur) if payeur else None},
                reimbursement=reimbursement,
            ))

        result.meta = {"sheet": "Suivi", "rows_read": len(result.transactions)}
        return result


parser = BdaExcelParser()
