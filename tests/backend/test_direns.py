"""Tests — module DirENS (export Excel, lignes = catégories, aucun mapping)."""
import io
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

from openpyxl import load_workbook


def _fy(client, name="2025-2026", start="2025-09-01"):
    r = client.post("/api/budget/fiscal-years", json={"name": name, "start_date": start})
    assert r.status_code == 201, r.text
    return r.json()


def _cat(client, name):
    return client.post("/api/categories/", json={"name": name}).json()


def _internal(client, name):
    return client.post("/api/entities/", json={"name": name, "type": "internal"}).json()


def _external(client, name="Fournisseur"):
    return client.post("/api/entities/", json={"name": name, "type": "external"}).json()


def _row_with_label(ws, label, max_row=60):
    """Numéro de ligne dont la colonne A vaut `label` (None si absente)."""
    for r in range(1, max_row + 1):
        if ws.cell(row=r, column=1).value == label:
            return r
    return None


# ─── Export : structure ───────────────────────────────────────────────────

def test_export_returns_xlsx(client):
    fy = _fy(client)
    r = client.get(f"/api/direns/export?bilan_fiscal_year_id={fy['id']}")
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers["content-type"]


def test_export_is_valid_workbook_three_sheets(client):
    fy = _fy(client)
    r = client.get(f"/api/direns/export?bilan_fiscal_year_id={fy['id']}")
    wb = load_workbook(io.BytesIO(r.content))
    assert len(wb.sheetnames) == 3


def test_export_writes_club_names_row5(client):
    _internal(client, "GastronomINE")
    fy = _fy(client)
    r = client.get(f"/api/direns/export?bilan_fiscal_year_id={fy['id']}")
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.worksheets[0]
    row5 = [ws.cell(row=5, column=c).value for c in range(2, 7)]
    assert "GastronomINE" in row5


def test_export_assoc_name_written(client):
    fy = _fy(client)
    r = client.get(f"/api/direns/export?bilan_fiscal_year_id={fy['id']}&assoc_name=Mon%20Asso")
    wb = load_workbook(io.BytesIO(r.content))
    assert wb.worksheets[0]["A3"].value == "Mon Asso"


def test_export_unknown_fy_returns_404(client):
    r = client.get("/api/direns/export?bilan_fiscal_year_id=9999")
    assert r.status_code == 404, r.text


# ─── Export : lignes = catégories ──────────────────────────────────────────

def test_expense_category_appears_as_row(client):
    ext = _external(client)
    club = _internal(client, "Club")
    cat = _cat(client, "Nourriture")
    fy = _fy(client)
    client.post("/api/transactions/", json={
        "date": "2025-10-01", "label": "Repas", "amount": 5000,  # 50 €
        "from_entity_id": club["id"], "to_entity_id": ext["id"], "category_id": cat["id"],
    })
    r = client.get(f"/api/direns/export?bilan_fiscal_year_id={fy['id']}")
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.worksheets[0]
    row = _row_with_label(ws, "Nourriture")
    assert row is not None, "la catégorie devrait apparaître en ligne"
    assert ws.cell(row=row, column=2).value == 50.0  # colonne B = premier club


def test_two_categories_two_rows(client):
    ext = _external(client)
    club = _internal(client, "Club")
    c1 = _cat(client, "Nourriture")
    c2 = _cat(client, "Transport")
    fy = _fy(client)
    client.post("/api/transactions/", json={
        "date": "2025-10-01", "label": "Repas", "amount": 5000,
        "from_entity_id": club["id"], "to_entity_id": ext["id"], "category_id": c1["id"],
    })
    client.post("/api/transactions/", json={
        "date": "2025-10-02", "label": "Train", "amount": 3000,
        "from_entity_id": club["id"], "to_entity_id": ext["id"], "category_id": c2["id"],
    })
    r = client.get(f"/api/direns/export?bilan_fiscal_year_id={fy['id']}")
    ws = load_workbook(io.BytesIO(r.content)).worksheets[0]
    assert ws.cell(row=_row_with_label(ws, "Nourriture"), column=2).value == 50.0
    assert ws.cell(row=_row_with_label(ws, "Transport"), column=2).value == 30.0


def test_uncategorized_expense_grouped(client):
    ext = _external(client)
    club = _internal(client, "Club")
    fy = _fy(client)
    client.post("/api/transactions/", json={
        "date": "2025-10-01", "label": "Sans cat", "amount": 1200,
        "from_entity_id": club["id"], "to_entity_id": ext["id"],
    })
    r = client.get(f"/api/direns/export?bilan_fiscal_year_id={fy['id']}")
    ws = load_workbook(io.BytesIO(r.content)).worksheets[0]
    row = _row_with_label(ws, "Non catégorisé")
    assert row is not None
    assert ws.cell(row=row, column=2).value == 12.0


def test_income_category_in_financing_section(client):
    ext = _external(client)
    club = _internal(client, "Club")
    cat = _cat(client, "Subventions")
    fy = _fy(client)
    # Recette : from externe -> to club
    client.post("/api/transactions/", json={
        "date": "2025-10-01", "label": "Subvention", "amount": 20000,  # 200 €
        "from_entity_id": ext["id"], "to_entity_id": club["id"], "category_id": cat["id"],
    })
    r = client.get(f"/api/direns/export?bilan_fiscal_year_id={fy['id']}")
    ws = load_workbook(io.BytesIO(r.content)).worksheets[0]
    row = _row_with_label(ws, "Subventions")
    assert row is not None
    assert ws.cell(row=row, column=2).value == 200.0


def test_total_label_present(client):
    ext = _external(client)
    club = _internal(client, "Club")
    cat = _cat(client, "Nourriture")
    fy = _fy(client)
    client.post("/api/transactions/", json={
        "date": "2025-10-01", "label": "Repas", "amount": 5000,
        "from_entity_id": club["id"], "to_entity_id": ext["id"], "category_id": cat["id"],
    })
    r = client.get(f"/api/direns/export?bilan_fiscal_year_id={fy['id']}")
    ws = load_workbook(io.BytesIO(r.content)).worksheets[0]
    assert _row_with_label(ws, "TOTAL DEPENSES REELLES") is not None


def test_budget_sheet_category_row(client):
    club = _internal(client, "Club")
    cat = _cat(client, "Matériel")
    fy = _fy(client)
    client.post(f"/api/budget/fiscal-years/{fy['id']}/allocations", json={
        "entity_id": club["id"], "category_id": cat["id"], "direction": "expense", "amount": 25000,  # 250 €
    })
    r = client.get(
        f"/api/direns/export?bilan_fiscal_year_id={fy['id']}&budget_fiscal_year_id={fy['id']}"
    )
    ws2 = load_workbook(io.BytesIO(r.content)).worksheets[1]
    row = _row_with_label(ws2, "Matériel")
    assert row is not None
    assert ws2.cell(row=row, column=2).value == 250.0


def test_parent_is_bold_header_children_indented(client):
    ext = _external(client)
    club = _internal(client, "Club")
    parent = _cat(client, "Poles")
    child = client.post("/api/categories/", json={"name": "Théâtre", "parent_id": parent["id"]}).json()
    fy = _fy(client)
    client.post("/api/transactions/", json={
        "date": "2025-10-01", "label": "Spectacle", "amount": 8000,  # 80 €
        "from_entity_id": club["id"], "to_entity_id": ext["id"], "category_id": child["id"],
    })
    r = client.get(f"/api/direns/export?bilan_fiscal_year_id={fy['id']}")
    ws = load_workbook(io.BytesIO(r.content)).worksheets[0]
    pr = _row_with_label(ws, "Poles")
    cr = _row_with_label(ws, "Théâtre")
    assert pr is not None and cr is not None
    # Parent = en-tête gras sans valeur ; enfant = indenté avec valeur.
    assert ws.cell(row=pr, column=1).font.bold is True
    assert ws.cell(row=pr, column=2).value in (None, 0)
    assert ws.cell(row=cr, column=2).value == 80.0
    assert (ws.cell(row=cr, column=1).alignment.indent or 0) >= 1


def test_inactive_club_excluded_from_columns(client):
    ext = _external(client)
    active = _internal(client, "Actif")
    _internal(client, "Inactif")  # aucun mouvement
    cat = _cat(client, "Nourriture")
    fy = _fy(client)
    client.post("/api/transactions/", json={
        "date": "2025-10-01", "label": "Repas", "amount": 5000,
        "from_entity_id": active["id"], "to_entity_id": ext["id"], "category_id": cat["id"],
    })
    r = client.get(f"/api/direns/export?bilan_fiscal_year_id={fy['id']}")
    ws = load_workbook(io.BytesIO(r.content)).worksheets[0]
    headers = [ws.cell(row=5, column=c).value for c in range(2, 8)]
    assert "Actif" in headers
    assert "Inactif" not in headers


def test_year_label_derived_from_dates(client):
    fy = client.post("/api/budget/fiscal-years", json={"name": "Mandat Truc", "start_date": "2024-09-01"}).json()
    client.post(f"/api/budget/fiscal-years/{fy['id']}/close", json={"end_date": "2025-08-31"})
    r = client.get(f"/api/direns/export?bilan_fiscal_year_id={fy['id']}")
    ws = load_workbook(io.BytesIO(r.content)).worksheets[0]
    assert ws["A1"].value == "BILAN FINANCIER 2024-2025"


def test_tresorerie_estimated_filled(client):
    club = _internal(client, "Club")
    fy = _fy(client)
    r = client.get(f"/api/direns/export?bilan_fiscal_year_id={fy['id']}")
    ws = load_workbook(io.BytesIO(r.content)).worksheets[0]
    row = _row_with_label(ws, "SOLDE TRESORERIE (A date)")
    assert row is not None
    assert isinstance(ws.cell(row=row, column=2).value, (int, float))


def test_bank_balance_non_derivable_is_placeholder(client):
    """La ligne compte bancaire (à date) n'est pas déductible -> placeholder, pas une valeur."""
    club = _internal(client, "Club")
    fy = _fy(client)
    r = client.get(f"/api/direns/export?bilan_fiscal_year_id={fy['id']}")
    ws = load_workbook(io.BytesIO(r.content)).worksheets[0]
    row = _row_with_label(ws, "SOLDE COMPTE BANCAIRE (A date)")
    assert row is not None
    assert ws.cell(row=row, column=2).value == "à compléter"


# ─── Règle unique : net par catégorie (net > 0 → financement, net < 0 → dépense) ──

def test_associated_income_deducted_from_expense(client):
    """Une recette propre (billetterie) réduit la dépense de la même catégorie/club."""
    ext = _external(client)
    club = _internal(client, "Club")
    cat = _cat(client, "Soirée")
    fy = _fy(client)
    # Dépense 800 €
    client.post("/api/transactions/", json={
        "date": "2025-10-01", "label": "Coûts soirée", "amount": 80000,
        "from_entity_id": club["id"], "to_entity_id": ext["id"], "category_id": cat["id"],
    })
    # Recette propre (billetterie) 600 € -> dépense nette = 200 €
    client.post("/api/transactions/", json={
        "date": "2025-10-02", "label": "Billetterie", "amount": 60000,
        "from_entity_id": ext["id"], "to_entity_id": club["id"], "category_id": cat["id"],
    })
    r = client.get(f"/api/direns/export?bilan_fiscal_year_id={fy['id']}")
    ws = load_workbook(io.BytesIO(r.content)).worksheets[0]
    row = _row_with_label(ws, "Soirée")
    assert row is not None
    assert ws.cell(row=row, column=2).value == 200.0
    # Aucune subvention -> total financement nul.
    fin = _row_with_label(ws, "TOTAL FINANCEMENT RECU 2025-2026")
    assert fin is not None
    assert ws.cell(row=fin, column=2).value in (0, None)


def test_income_and_expense_on_separate_categories_not_compensated(client):
    """Recette et dépense sur des catégories DIFFÉRENTES ne se compensent pas : la dépense
    reste en dépense, la recette pure part en financement sous le nom de sa catégorie."""
    ext = _external(client)
    club = _internal(client, "Club")
    dep = _cat(client, "Nourriture")
    subv = _cat(client, "Subvention DirENS")
    fy = _fy(client)
    client.post("/api/transactions/", json={
        "date": "2025-10-01", "label": "Repas", "amount": 50000,  # 500 € dépense
        "from_entity_id": club["id"], "to_entity_id": ext["id"], "category_id": dep["id"],
    })
    client.post("/api/transactions/", json={
        "date": "2025-10-02", "label": "Subvention", "amount": 30000,  # 300 € recette
        "from_entity_id": ext["id"], "to_entity_id": club["id"], "category_id": subv["id"],
    })
    r = client.get(f"/api/direns/export?bilan_fiscal_year_id={fy['id']}")
    ws = load_workbook(io.BytesIO(r.content)).worksheets[0]
    # La dépense (catégorie Nourriture) reste à 500 €.
    drow = _row_with_label(ws, "Nourriture")
    assert drow is not None
    assert ws.cell(row=drow, column=2).value == 500.0
    # La recette (catégorie Subvention DirENS) apparaît en financement à 300 €.
    srow = _row_with_label(ws, "Subvention DirENS")
    assert srow is not None
    assert ws.cell(row=srow, column=2).value == 300.0


def test_income_and_expense_same_category_are_netted(client):
    """Règle unique : recette et dépense d'une MÊME catégorie se compensent. Si la recette
    l'emporte, la catégorie apparaît en financement pour son solde NET (et nulle part en
    dépense) — y compris une « subvention » imputée sur une catégorie qui a aussi des frais."""
    ext = _external(client)
    club = _internal(client, "Club")
    cat = _cat(client, "Campagne don")
    fy = _fy(client)
    client.post("/api/transactions/", json={
        "date": "2025-10-01", "label": "Frais", "amount": 5000,  # 50 € dépense
        "from_entity_id": club["id"], "to_entity_id": ext["id"], "category_id": cat["id"],
    })
    client.post("/api/transactions/", json={
        "date": "2025-10-02", "label": "Dons reçus", "amount": 30000,  # 300 € recette
        "from_entity_id": ext["id"], "to_entity_id": club["id"], "category_id": cat["id"],
    })
    r = client.get(f"/api/direns/export?bilan_fiscal_year_id={fy['id']}")
    ws = load_workbook(io.BytesIO(r.content)).worksheets[0]
    total_row = _row_with_label(ws, "TOTAL DEPENSES REELLES")
    # Aucune ligne de dépense pour la catégorie : la dépense est compensée.
    assert [r for r in range(1, total_row) if ws.cell(row=r, column=1).value == "Campagne don"] == []
    # Solde net (300 − 50 = 250 €) en financement, sous le nom de la catégorie.
    frow = _row_with_label_after(ws, "Campagne don", total_row)
    assert frow is not None
    assert ws.cell(row=frow, column=2).value == 250.0
    # Garde-fou : l'agrégat « Recettes propres » de l'ancien système n'existe plus.
    assert _row_with_label(ws, "Recettes propres") is None


def _row_with_label_after(ws, label, after_row, max_row=60):
    """Numéro de ligne dont la colonne A vaut `label`, cherché APRÈS `after_row`."""
    for r in range(after_row + 1, max_row + 1):
        if ws.cell(row=r, column=1).value == label:
            return r
    return None


def test_pure_income_keeps_its_name_in_financing(client):
    """Une recette pure (aucune dépense sur la même catégorie) garde son NOM en financement
    sous le total (et il n'existe plus de ligne agrégée « Recettes propres »)."""
    ext = _external(client)
    club = _internal(client, "Club")
    subv = _cat(client, "Aide région")
    fy = _fy(client)
    client.post("/api/transactions/", json={
        "date": "2025-10-02", "label": "Subvention région", "amount": 40000,  # 400 €
        "from_entity_id": ext["id"], "to_entity_id": club["id"], "category_id": subv["id"],
    })
    r = client.get(f"/api/direns/export?bilan_fiscal_year_id={fy['id']}")
    ws = load_workbook(io.BytesIO(r.content)).worksheets[0]
    total_row = _row_with_label(ws, "TOTAL DEPENSES REELLES")
    srow = _row_with_label_after(ws, "Aide région", total_row)
    assert srow is not None
    assert ws.cell(row=srow, column=2).value == 400.0
    assert _row_with_label(ws, "Recettes propres") is None


def test_totals_are_formulas_with_cached_values(client):
    """Les totaux restent des FORMULES (=SUM...) ET portent leur valeur calculée en cache,
    pour s'afficher dans tout lecteur sans recalcul (avant : <v> vide -> cellules vides)."""
    ext = _external(client)
    a = _internal(client, "Alpha")
    b = _internal(client, "Beta")
    nour = _cat(client, "Nourriture")
    mat = _cat(client, "Materiel")
    subv = _cat(client, "Subventions")
    fy = _fy(client)

    def tx(frm, to, amount, cat):
        client.post("/api/transactions/", json={
            "date": "2025-10-01", "label": "x", "amount": amount,
            "from_entity_id": frm, "to_entity_id": to, "category_id": cat,
        })
    tx(a["id"], ext["id"], 10000, nour["id"])   # Alpha nourriture 100 €
    tx(b["id"], ext["id"], 5000, nour["id"])    # Beta nourriture 50 €
    tx(a["id"], ext["id"], 3000, mat["id"])     # Alpha matériel 30 €
    tx(ext["id"], a["id"], 20000, subv["id"])   # subvention 200 € -> Alpha

    r = client.get(f"/api/direns/export?bilan_fiscal_year_id={fy['id']}")
    data = r.content
    fsheet = load_workbook(io.BytesIO(data)).worksheets[0]                 # formules
    vsheet = load_workbook(io.BytesIO(data), data_only=True).worksheets[0]  # valeurs en cache

    cols = {fsheet.cell(row=5, column=c).value: c for c in range(2, 7)
            if fsheet.cell(row=5, column=c).value}
    ca, cb = cols["Alpha"], cols["Beta"]

    # Somme horizontale (colonne TOTAL) d'une ligne catégorie = 100 + 50.
    rn = _row_with_label(fsheet, "Nourriture")
    assert str(fsheet.cell(row=rn, column=6).value).startswith("=SUM")
    assert vsheet.cell(row=rn, column=6).value == 150.0

    # Ligne TOTAL : sommes verticales par club + total général.
    rt = _row_with_label(fsheet, "TOTAL DEPENSES REELLES")
    assert str(fsheet.cell(row=rt, column=6).value).startswith("=SUM")
    assert vsheet.cell(row=rt, column=ca).value == 130.0   # Alpha 100 + 30
    assert vsheet.cell(row=rt, column=cb).value == 50.0    # Beta 50
    assert vsheet.cell(row=rt, column=6).value == 180.0    # total général

    # Total financement + rappel de la dépense réelle en bas (formule + valeur).
    rf = _row_with_label(fsheet, "TOTAL FINANCEMENT RECU 2025-2026")
    assert str(fsheet.cell(row=rf, column=2).value).startswith("=SUM")
    assert vsheet.cell(row=rf, column=2).value == 200.0
    rr = _row_with_label(fsheet, "TOTAL DEPENSES REELLES 2025-2026 (cf tableau ci-dessus)")
    assert str(fsheet.cell(row=rr, column=2).value).startswith("=")
    assert vsheet.cell(row=rr, column=2).value == 180.0


def test_many_categories_expand_rows(client):
    """Plus de 26 catégories : le bloc s'agrandit, toutes les lignes présentes."""
    ext = _external(client)
    club = _internal(client, "Club")
    fy = _fy(client)
    names = [f"Categorie {i:02d}" for i in range(30)]
    for i, nm in enumerate(names):
        c = _cat(client, nm)
        client.post("/api/transactions/", json={
            "date": "2025-10-01", "label": nm, "amount": 100 * (i + 1),
            "from_entity_id": club["id"], "to_entity_id": ext["id"], "category_id": c["id"],
        })
    r = client.get(f"/api/direns/export?bilan_fiscal_year_id={fy['id']}")
    ws = load_workbook(io.BytesIO(r.content)).worksheets[0]
    for nm in names:
        assert _row_with_label(ws, nm) is not None, f"{nm} manquante"
    # Le total reste cohérent (présent sous les catégories).
    assert _row_with_label(ws, "TOTAL DEPENSES REELLES") is not None
